"""
Excel Generator Module for LSS Invoice Automation
Creates Excel files in LawSync format
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel output in LawSync format"""
    
    def __init__(self, output_path, headers):
        self.output_path = output_path
        self.headers = headers
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Line Items"
        self.current_row = 1
    
    def format_headers(self):
        """Apply formatting to header row"""
        # Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_side = Side(style="thin", color="000000")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Write and format headers
        for col_idx, header in enumerate(self.headers, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set row height
        self.worksheet.row_dimensions[1].height = 30
        
        logger.debug("Headers formatted")
    
    def populate_data_rows(self, invoice_data_list):
        """
        Populate worksheet with line item data
        
        Args:
            invoice_data_list: List of parsed invoice data dictionaries
        """
        row_idx = 2  # Start after header
        
        for invoice_data in invoice_data_list:
            if not invoice_data:
                continue

            header = invoice_data.get("header", {})
            line_items = invoice_data.get("line_items", [])
            totals = invoice_data.get("totals", {})

            invoice_number = header.get("invoice_number", "")
            client = header.get("client", "")
            invoice_date = header.get("invoice_date", "")
            matter_name = header.get("matter_name", "")
            appeal_expiry = header.get("appeal_expiry", "")
            total_submitted = totals.get("submitted", 0)

            for item in line_items:
                # Only export line items that have a reduction
                reduced_amount = item.get("reduced_amount", 0)
                if reduced_amount <= 0:
                    continue

                description = item.get("description", "")
                audit_reason = item.get("audit_reason", "")

                # Narrative = description + audit reason block
                narrative = description
                if audit_reason:
                    narrative += f"\nAudit Reason : {audit_reason}"

                # Item Type: 'Fees' when timekeeper initials present, else 'Expenses'
                tk_initials = item.get("timekeeper_id", "").strip()
                item_type_display = "Fees" if tk_initials else "Expenses"

                # 15-column LawSync layout
                values = [
                    invoice_number,              # A  Invoice Number
                    client,                      # B  Client (insurance company)
                    invoice_date,                # C  Invoice Date
                    item.get("timekeeper", ""),  # D  Working Timekeeper
                    matter_name,                 # E  Matter Name
                    item.get("date", ""),        # F  Date of Item
                    appeal_expiry,               # G  Appeal Expiry (Finalized + 30 days)
                    item_type_display,           # H  Item Type
                    item.get("units", 0),        # I  UNITS
                    item.get("rate", 0),         # J  RATE
                    item.get("amount", 0),       # K  AMOUNT
                    reduced_amount,              # L  Reduced Amount
                    total_submitted,             # M  Total Invoice Amount
                    narrative,                   # N  Narrative
                    audit_reason,                # O  Adjustment Reason
                ]

                # Write values to row
                for col_idx, value in enumerate(values, start=1):
                    cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value

                    # Currency: RATE (J=10), AMOUNT (K=11), Reduced Amount (L=12), Total (M=13)
                    if col_idx in [10, 11, 12, 13]:
                        cell.number_format = '$#,##0.00'

                    # Units (I=9)
                    if col_idx == 9:
                        cell.number_format = '0.00'

                    # Center-align dates and units
                    if col_idx in [3, 6, 7, 9]:  # Invoice Date, Date of Item, Appeal Expiry, UNITS
                        cell.alignment = Alignment(horizontal="center")

                row_idx += 1
        
        total_rows = row_idx - 2
        logger.info(f"Populated {total_rows} data rows")
        return total_rows
    
    def apply_column_widths(self):
        """Auto-adjust column widths for readability"""
        # Column widths for 15-column LawSync layout
        column_widths = {
            'A': 15,  # Invoice Number
            'B': 30,  # Client
            'C': 14,  # Invoice Date
            'D': 25,  # Working Timekeeper
            'E': 55,  # Matter Name
            'F': 14,  # Date of Item
            'G': 14,  # Appeal Expiry
            'H': 12,  # Item Type
            'I': 10,  # UNITS
            'J': 12,  # RATE
            'K': 12,  # AMOUNT
            'L': 15,  # Reduced Amount
            'M': 20,  # Total Invoice Amount
            'N': 60,  # Narrative
            'O': 50,  # Adjustment Reason
        }

        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width

        logger.debug("Column widths applied")

    
    def save(self):
        """Save the workbook to file"""
        try:
            self.workbook.save(self.output_path)
            logger.info(f"Excel file saved: {self.output_path}")
            return True
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}", exc_info=True)
            return False


def create_lawsync_excel(invoice_data_list, output_path, headers):
    """
    Main function to create LawSync format Excel file
    
    Args:
        invoice_data_list: List of parsed invoice data dictionaries
        output_path: Path where Excel file should be saved
        headers: List of column headers
    
    Returns:
        True if successful, False otherwise
    """
    try:
        generator = ExcelGenerator(output_path, headers)
        
        # Format headers
        generator.format_headers()
        
        # Populate data
        total_rows = generator.populate_data_rows(invoice_data_list)
        
        # Apply column widths
        generator.apply_column_widths()
        
        # Save file
        return generator.save()
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}", exc_info=True)
        return False


def append_to_master_tracker(invoice_data_list, tracker_path, headers, formula_template):
    """
    Append summary data to the Master Tracker Excel file
    
    Args:
        invoice_data_list: List of parsed invoice data dictionaries
        tracker_path: Path to the Master Tracker Excel file
        headers: List of column headers for the tracker
        formula_template: Formula for the 'ExistInMaster' column
    """
    try:
        import os
        from openpyxl import load_workbook, Workbook
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(tracker_path), exist_ok=True)
        
        if os.path.exists(tracker_path):
            wb = load_workbook(tracker_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header
        
        start_row = ws.max_row + 1
        current_row = start_row
        
        for data in invoice_data_list:
            if not data:
                continue
            
            header = data.get("header", {})
            totals = data.get("totals", {})
            
            # Map internal keys to headers
            # headers: ['Email Date', 'Invoice', 'Invoice Date', 'Invoice Amount', 
            #           'Total Adjustment', 'Approved Total', 'ExistInMaster', 'Client', 'Appeal Exp']
            
            row_data = {
                'Email Date': data.get('email_date', ''),
                'Invoice': header.get('invoice_number', ''),
                'Invoice Date': header.get('invoice_date', ''),
                'Invoice Amount': totals.get('submitted', 0),
                'Total Adjustment': totals.get('reduction', 0),
                'Approved Total': totals.get('payable', 0),
                'ExistInMaster': formula_template.format(row=current_row),
                'Client': data.get('client', ''),
                'Appeal Exp': header.get('finalized_date', '')
            }
            
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                val = row_data.get(h, '')
                
                # Handle formula
                if h == 'ExistInMaster':
                    cell.value = val
                else:
                    cell.value = val
                
                # Formatting
                if h in ['Invoice Amount', 'Total Adjustment', 'Approved Total']:
                    cell.number_format = '$#,##0.00'
                elif h in ['Email Date', 'Invoice Date', 'Appeal Exp']:
                    cell.alignment = Alignment(horizontal="center")
            
            current_row += 1
            
        wb.save(tracker_path)
        logger.info(f"Appended {len(invoice_data_list)} records to Master Tracker: {tracker_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error appending to Master Tracker: {str(e)}", exc_info=True)
        return False
